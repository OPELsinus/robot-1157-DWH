import datetime
import os
import re
import shutil
import sys
from contextlib import suppress
from pathlib import Path

import openpyxl

from config import owa_username, owa_password, local_path, saving_path, download_path, smtp_host, smtp_author, chat_id, bot_token, logger, calendar_path

import psycopg2
import csv

from openpyxl import load_workbook
import time

import pandas as pd

from tools import update_credentials, send_message_by_smtp, send_message_to_tg


def sql_request(start_date, end_date):
    conn = psycopg2.connect(dbname='adb', host='172.16.10.22', port='5432',
                            user='rpa_robot', password='Qaz123123+')

    cur = conn.cursor()

    cur.execute(f"""select dpl.article as "Артикул", pg1.name_group_level5 as "Подгруппа", dpl.name_wares as "Наименование товара", dpl.extdesc28 as "Ценовой сегмент", ds.sale_obj_name as "Филиал", fwb.source_product_id as "Код товара", fwb.source_store_id as "Код филиала", sum(quantity_warehouse_acc) as "Учётные остатки", sum(quantity_warehouse) as "Фактические остатки", sum(quantity_free) as "Свободные остатки", sum(quantity_problem) as "Кол-во проблемных"
    from dwh_data.fact_wares_bal fwb
    left join dwh_data.dim_store ds on ds.source_store_id = fwb.source_store_id and current_date between ds.datestart and ds.dateend 
    left join dwh_data.dim_prod_group pg1 on pg1.source_prod_id = fwb.source_product_id and  fwb.rep_date between pg1.datestart and pg1.dateend
    left join dwh_data.dim_product_list dpl on dpl.code_wares = fwb.source_product_id and current_date between dpl.datestart and dpl.dateend 
    where fwb.rep_date = '{end_date}' and pg1.code_group_level5 in ('1398', '1399', '1417', '1437', '3951', '1745', '1746', '1747', '1754')
    and ds.sale_obj_name = ds.sale_obj_name and dpl.extdesc28 = 'эконом'
    group by fwb.source_product_id, pg1.name_group_level5, dpl.name_wares, dpl.extdesc28, dpl.article, ds.sale_obj_name, fwb.source_store_id
    order by dpl.name_wares;
    """)

    print((f"""select dpl.article as "Артикул", pg1.name_group_level5 as "Подгруппа", dpl.name_wares as "Наименование товара", dpl.extdesc28 as "Ценовой сегмент", ds.sale_obj_name as "Филиал", fwb.source_product_id as "Код товара", fwb.source_store_id as "Код филиала", sum(quantity_warehouse_acc) as "Учётные остатки", sum(quantity_warehouse) as "Фактические остатки", sum(quantity_free) as "Свободные остатки", sum(quantity_problem) as "Кол-во проблемных"
    from dwh_data.fact_wares_bal fwb
    left join dwh_data.dim_store ds on ds.source_store_id = fwb.source_store_id and current_date between ds.datestart and ds.dateend 
    left join dwh_data.dim_prod_group pg1 on pg1.source_prod_id = fwb.source_product_id and  fwb.rep_date between pg1.datestart and pg1.dateend
    left join dwh_data.dim_product_list dpl on dpl.code_wares = fwb.source_product_id and current_date between dpl.datestart and dpl.dateend 
    where fwb.rep_date = '{end_date}' and pg1.code_group_level5 in ('1398', '1399', '1417', '1437', '3951', '1745', '1746', '1747', '1754')
    and ds.sale_obj_name = ds.sale_obj_name and dpl.extdesc28 = 'эконом'
    group by fwb.source_product_id, pg1.name_group_level5, dpl.name_wares, dpl.extdesc28, dpl.article, ds.sale_obj_name, fwb.source_store_id
    order by dpl.name_wares;
    """))

    rows = cur.fetchall()

    with open(os.path.join(saving_path, 'all.csv'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        for row in rows:
            writer.writerow(row)

    df = pd.DataFrame(rows)
    df.to_excel(os.path.join(saving_path, 'all1.xlsx'))

    cur.close()
    conn.close()


def dividing_into_single_reports(saving_path_, date_):
    try:
        os.makedirs(os.path.join(saving_path_, f'Splitted'))
    except:
        pass

    df = pd.read_csv(os.path.join(saving_path_, 'all.csv'), header=None)
    # df = df.drop(df.columns[11:], axis=1)'
    df.columns = ["Артикул", "Подгруппа", "Наименование товара", "Ценовой сегмент", "Филиал", "Код товара", "Код филиала", "Учётные остатки", "Фактические остатки", "Свободные остатки", "Кол-во проблемных"]

    for i in range(len(df)):

        df['Кол-во проблемных'].iloc[i] = df['Фактические остатки'].iloc[i]
        df['Фактические остатки'].iloc[i] /= 1000

    # * ---------- МАСЛО ----------

    df1 = df[df['Подгруппа'].str.lower().str.contains('масло')].copy()

    for i in range(len(df1)):

        df1['Наименование товара'] = df1['Наименование товара'].str.replace(',', '.')

        multiplier = [float(s) for s in re.findall(r'[-+]?\d*\.\d+|\d+', df1['Наименование товара'].iloc[i])]

        multiplier_ = float(max(multiplier))

        if 'мл ' in df1['Наименование товара'].iloc[i].lower() or str(df1['Наименование товара'].iloc[i].lower())[-2:] == 'мл':
            multiplier_ /= 1000
        if multiplier_ >= 1000:
            multiplier_ /= 1000

        # df1['Кол-во проблемных'].iloc[i] = df1['Фактические остатки'].iloc[i] * 1000
        df1['Фактические остатки'].iloc[i] *= multiplier_ * 1000

        # df1['Свободные остатки'].iloc[i] = 'lol'

    # * ---------- Занесение изменений в датафрейм ----------

    for i in df1.index:
        if i < df1['Фактические остатки'].index[0] or i > df1['Фактические остатки'].index[-1]:
            pass
        else:
            idx_in_df1 = i - df1['Фактические остатки'].index[0]
            if idx_in_df1 >= 0 and idx_in_df1 < len(df1['Фактические остатки']):
                df.loc[i, 'Фактические остатки'] = df1['Фактические остатки'].iloc[idx_in_df1]
                # df.loc[i, 'Свободные остатки'] = df1['Свободные остатки'].iloc[idx_in_df1]
                df.loc[i, 'Кол-во проблемных'] = df1['Кол-во проблемных'].iloc[idx_in_df1]

    # * ---------- СОЛЬ ----------

    df1 = df[df['Подгруппа'].str.lower().str.contains('соль')].copy()

    for i in range(len(df1)):

        df1['Наименование товара'] = df1['Наименование товара'].str.replace(',', '.')

        # df1['Кол-во проблемных'].iloc[i] = df1['Фактические остатки'].iloc[i] * 1000
        df1['Фактические остатки'].iloc[i] *= 1000

        # df1['Свободные остатки'].iloc[i] = 'lol'

    # * ---------- Занесение изменений в датафрейм ----------

    for i in df1.index:
        if i < df1['Фактические остатки'].index[0] or i > df1['Фактические остатки'].index[-1]:
            print('skip', i, df1['Фактические остатки'].index[0], df1['Фактические остатки'].index[-1])
            pass
        else:
            idx_in_df1 = i - df1['Фактические остатки'].index[0]
            if 0 <= idx_in_df1 < len(df1['Фактические остатки']):
                df.loc[i, 'Фактические остатки'] = df1['Фактические остатки'].iloc[idx_in_df1]
                # df.loc[i, 'Свободные остатки'] = df1['Свободные остатки'].iloc[idx_in_df1]
                df.loc[i, 'Кол-во проблемных'] = df1['Кол-во проблемных'].iloc[idx_in_df1]

    print('Saving into files')

    for i in sorted(df['Филиал'].unique()):

        df1 = df.iloc[df[df['Филиал'] == i].index]

        df1.columns = ["Артикул", "Подгруппа", "Наименование товара", "Ценовой сегмент", "Филиал", "Код товара", "Код филиала", "Учётные остатки", "Фактические остатки", "Свободные остатки", "Факт остатки - ОРИГИНАЛ"]

        df1.to_excel(os.path.join(saving_path_, f'Splitted\\{i}_{date_}.xlsx'), index=False)
        time.sleep(1)

        book = load_workbook(os.path.join(saving_path_, f'Splitted\\{i}_{date_}.xlsx'))

        worksheet = book.active

        column_widths = []
        for j, column in enumerate(df1.columns):
            column_widths.append(max(df1[column].astype(str).map(len).max(), len(column)) * 1.1)

        for j, width in enumerate(column_widths):
            worksheet.column_dimensions[worksheet.cell(row=1, column=j + 1).column_letter].width = width

        book.save(os.path.join(saving_path_, f'Splitted\\{i}_{date_}.xlsx'))


def archive_files(prev_date):

    folder_path = os.path.join(saving_path, f'Splitted')

    try:
        os.makedirs(os.path.join(saving_path, f'Splitted1'))
    except:
        pass
    destination_folder = os.path.join(saving_path, f'Splitted1')

    zip_file_name = f'Все филиалы - 1157 за {prev_date}'
    zip_file_path = os.path.join(destination_folder, zip_file_name)

    shutil.make_archive(zip_file_path, 'zip', folder_path)

    return zip_file_path


def is_today_start():

    calendar = pd.read_excel(calendar_path)

    today_ = datetime.datetime.now().strftime('%d.%m.%y')

    cur_day_index = calendar[calendar['Day'] == today_]['Type'].index[0]
    cur_day_type = calendar[calendar['Day'] == today_]['Type'].iloc[0]

    count = 0
    day_ = None
    found = False

    for i in range(1, 31):

        try:
            day = int(calendar['Day'].iloc[cur_day_index + i].split('.')[0])
            print(calendar['Day'].iloc[cur_day_index + i], calendar['Weekday'].iloc[cur_day_index + i], calendar['Type'].iloc[cur_day_index + i])

        except:
            day = 1

        if day == 1:

            for j in range(1, 6):
                print(cur_day_index, i, j, cur_day_index + i - j)

                print('---', calendar['Day'].iloc[cur_day_index + i - j], calendar['Weekday'].iloc[cur_day_index + i - j], calendar['Type'].iloc[cur_day_index + i - j])

                if calendar['Type'].iloc[cur_day_index + i - j] == 'Working':
                    count += 1
                if count == 3:
                    found = True
                    day_ = calendar['Day'].iloc[cur_day_index + i - j]
                    break
        if found:
            break

    print(cur_day_index, cur_day_type)

    print(day_)

    if today_ == day_:  # * datetime.datetime.today().strftime('%d.%m.%y') == day_:
        return True
    else:
        return False


if __name__ == '__main__':

    if not is_today_start():
        logger.info(f'Not working day - {datetime.date.today()}')
        exit()

    update_credentials(Path(r'\\172.16.8.87\d'), owa_username, owa_password)

    with suppress(Exception):
        shutil.rmtree(os.path.join(saving_path, 'Splitted'))
    with suppress(Exception):
        shutil.rmtree(os.path.join(saving_path, 'Splitted1'))
    with suppress(Exception):
        Path.unlink(Path(os.path.join(saving_path, 'all.csv')))

    df = pd.read_excel(calendar_path)
    calendar = pd.read_excel(calendar_path)
    # curr_date = df['Day'].iloc[0]
    curr_date = datetime.datetime.now().strftime('%d.%m.%y')

    # today = datetime.date.today()
    # today = datetime.datetime(2023, 8, 2)
    # first_day_of_current_month = datetime.date(today.year, today.month, 1)
    #
    # if today.month == 12:
    #     last_day_of_current_month = datetime.date(today.year + 1, 1, 1) - datetime.timedelta(days=1)
    # else:
    #     last_day_of_current_month = datetime.date(today.year, today.month + 1, 1) - datetime.timedelta(days=1)

    today = datetime.date.today()
    first_day_of_current_month = today - datetime.timedelta(days=31)  # datetime.date(today.year, today.month, 1)
    last_day_of_current_month = today - datetime.timedelta(days=1)

    print(first_day_of_current_month, last_day_of_current_month)

    # print(curr_date)
    # day = int(curr_date.split('.')[0])
    # month = int(curr_date.split('.')[1])
    # year = int('20' + curr_date.split('.')[2])
    # # print(datetime.date(year, month, day))
    # prev_date = datetime.date(year, month, day)
    # prev_date = (prev_date - datetime.timedelta(days=1))
    # print(prev_date)

    sql_request(first_day_of_current_month, last_day_of_current_month)

    send_message_to_tg(bot_token=bot_token, message='Started dividing', chat_id=chat_id)

    dividing_into_single_reports(saving_path, last_day_of_current_month)

    filepath = archive_files(last_day_of_current_month)

    send_message_to_tg(bot_token=bot_token, message=f"Сегодня: {curr_date}\nОтрабатывал за: {datetime.date.today().strftime('%d.%m.%y')}\nДата создания zip файла:\n{time.ctime(os.path.getctime(filepath + '.zip'))}", chat_id=chat_id)

    send_message_by_smtp(smtp_host, to=['Abdykarim.D@magnum.kz', 'Mukhtarova@magnum.kz'], subject=f'Отчёт 1157 за {datetime.date.today().strftime("%d.%m.%Y")}',
                         body='Результаты в приложении', username=smtp_author,
                         attachments=[filepath + '.zip'])

    Path(filepath + '.zip').unlink()



